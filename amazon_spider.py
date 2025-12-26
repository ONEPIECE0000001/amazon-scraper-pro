import scrapy
from scrapy_playwright.page import PageMethod
import random
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re
import json
import asyncio
from playwright.async_api import async_playwright


class AmazonSpider(scrapy.Spider):
    name = "amazon_spider"
    allowed_domains = ["amazon.com"]
    custom_settings = {
        "PLAYWRIGHT_BROWSER_TYPE": "chromium",
        "PLAYWRIGHT_DEFAULT_NAVIGATION_TIMEOUT": 60000,
        "PLAYWRIGHT_LAUNCH_OPTIONS": {
            "headless": True,
            "timeout": 30000,
            "args": [
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--disable-web-security",
                "--disable-features=VizDisplayCompositor"
            ]
        }
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Amazon商品数据"
        self.worksheet.append([
            "商品ID", "商品名称", "价格", "销量", "评分", "评论数", 
            "品牌", "类别", "上架时间", "采集时间", "URL"
        ])
        
        # 配置参数
        self.min_wait_time = 3
        self.max_wait_time = 8
        self.scroll_wait_time = 3000
        self.retry_times = 3
        
        # 反爬策略
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15'
        ]

    def start_requests(self):
        # 示例：搜索"laptop"商品
        search_terms = getattr(self, 'search', 'laptop')
        urls = [
            f"https://www.amazon.com/s?k={search_terms}&page=1",
            f"https://www.amazon.com/s?k={search_terms}&page=2",
            # 可以添加更多页面URL
        ]
        
        for url in urls:
            yield scrapy.Request(
                url,
                meta={
                    "playwright": True,
                    "playwright_page_coroutines": [
                        PageMethod("wait_for_load_state", "networkidle"),
                        PageMethod("wait_for_selector", "[data-component-type='s-search-result']"),
                    ],
                    "playwright_context_kwargs": {
                        "user_agent": random.choice(self.user_agents)
                    }
                },
                callback=self.parse,
                errback=self.errback_close_page,
                headers={
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                    "Accept-Language": "en-US,en;q=0.5",
                    "Accept-Encoding": "gzip, deflate",
                    "Connection": "keep-alive",
                    "Upgrade-Insecure-Requests": "1",
                }
            )

    async def parse(self, response):
        self.logger.info(f"Parsing page: {response.url}")

        try:
            # 模拟随机停留时间，模拟人类行为
            await asyncio.sleep(random.uniform(self.min_wait_time, self.max_wait_time))

            # 获取所有商品项
            products = response.css("[data-component-type='s-search-result']")
            
            for product in products:
                try:
                    # 提取商品信息
                    product_id = product.css("::attr(data-asin)").get()
                    if not product_id:
                        continue
                        
                    # 商品名称
                    name = product.css("h2 a span::text").get()
                    if name:
                        name = name.strip()
                    
                    # 价格（可能有多种格式）
                    price_whole = product.css(".a-price-whole::text").get()
                    price_fraction = product.css(".a-price-fraction::text").get()
                    if price_whole:
                        price = f"${price_whole}{price_fraction or ''}"
                    else:
                        price = product.css(".a-offscreen::text").get()
                    
                    # 销量和评分
                    rating = product.css("span[aria-label]::attr(aria-label)").get()
                    review_count = product.css("span[aria-label*='ratings']::text").re(r'\d+')
                    review_count = review_count[0] if review_count else "0"
                    
                    # 如果没有直接的销量信息，尝试其他选择器
                    if review_count == "0":
                        review_count_alt = product.css("a[role='link'] span::text").re(r'\d+')
                        review_count = review_count_alt[0] if review_count_alt else "0"
                    
                    # 品牌（需要进入详情页获取）
                    brand = "N/A"
                    
                    # 构建商品详情页URL
                    product_url = product.css("h2 a::attr(href)").get()
                    if product_url:
                        product_url = response.urljoin(product_url)
                    else:
                        product_url = "N/A"
                    
                    # 数据验证和清理
                    if product_id and name:
                        # 清理数据
                        cleaned_name = name.strip() if name else ""
                        cleaned_price = price.strip() if price else "N/A"
                        cleaned_rating = rating.strip() if rating else "N/A"
                        cleaned_review_count = review_count.strip() if review_count else "0"

                        # 验证数据不为空
                        if cleaned_name:
                            self.worksheet.append([
                                product_id,
                                cleaned_name,
                                cleaned_price,
                                "N/A",  # 销量暂时设为N/A，因为需要从详情页获取
                                cleaned_rating,
                                cleaned_review_count,
                                brand,
                                "Electronics",  # 分类暂时设为Electronics
                                "N/A",  # 上架时间需要从详情页获取
                                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                product_url
                            ])
                            
                            # 请求商品详情页以获取更多信息
                            yield scrapy.Request(
                                url=product_url,
                                meta={
                                    "playwright": True,
                                    "playwright_page_coroutines": [
                                        PageMethod("wait_for_load_state", "networkidle"),
                                        PageMethod("wait_for_selector", "#productTitle"),
                                    ],
                                    "playwright_context_kwargs": {
                                        "user_agent": random.choice(self.user_agents)
                                    },
                                    "product_id": product_id
                                },
                                callback=self.parse_product_detail,
                                errback=self.errback_close_page,
                                headers={
                                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                                    "Accept-Language": "en-US,en;q=0.5",
                                    "Accept-Encoding": "gzip, deflate",
                                    "Connection": "keep-alive",
                                    "Upgrade-Insecure-Requests": "1",
                                }
                            )
                            
                except Exception as e:
                    self.logger.error(f"Error parsing product on {response.url}: {str(e)}")
                    continue

            # 处理分页
            next_page = response.css('a[aria-label="Next page"]::attr(href)').get()
            if next_page:
                next_page_url = response.urljoin(next_page)
                yield scrapy.Request(
                    next_page_url,
                    meta={
                        "playwright": True,
                        "playwright_page_coroutines": [
                            PageMethod("wait_for_load_state", "networkidle"),
                            PageMethod("wait_for_selector", "[data-component-type='s-search-result']"),
                        ],
                        "playwright_context_kwargs": {
                            "user_agent": random.choice(self.user_agents)
                        }
                    },
                    callback=self.parse,
                    errback=self.errback_close_page,
                    headers={
                        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                        "Accept-Language": "en-US,en;q=0.5",
                        "Accept-Encoding": "gzip, deflate",
                        "Connection": "keep-alive",
                        "Upgrade-Insecure-Requests": "1",
                    }
                )
        except Exception as e:
            self.logger.error(f"Error parsing page {response.url}: {str(e)}")

    async def parse_product_detail(self, response):
        """解析商品详情页"""
        product_id = response.meta.get("product_id")
        
        try:
            # 提取品牌
            brand = response.css("#bylineInfo::text").get()
            if not brand:
                brand = response.css("#brand ::text").get()
            
            # 提取销量
            bestseller_badge = response.css("span.badge-text::text").get()
            if not bestseller_badge:
                bestseller_badge = "N/A"
            
            # 提取上架时间
            date_info = response.css("#averageCustomerReviews span::text").getall()
            if date_info:
                # 尝试从评论信息中提取上架时间
                date_str = " ".join(date_info).strip()
            else:
                date_str = "N/A"
            
            # 提取分类
            category_elements = response.css("#wayfinding-breadcrumbs_feature_div ul li span a::text").getall()
            category = category_elements[-1] if category_elements else "Electronics"
            
            # 查找对应的商品行并更新详细信息
            for row in self.worksheet.iter_rows(min_row=2, max_row=self.worksheet.max_row):
                if row[0].value == product_id:  # 商品ID匹配
                    # 更新行数据
                    row[6].value = brand.strip() if brand else "N/A"
                    row[7].value = category
                    row[8].value = date_str
                    break
        except Exception as e:
            self.logger.error(f"Error parsing product detail {response.url}: {str(e)}")

    def close(self, reason):
        try:
            # 保存Excel文件，防止路径遍历
            current_time = datetime.now().strftime("%Y%m%d%H%M%S")
            filename = f"amazon_data_{current_time}.xlsx"
            # 验证文件名安全
            if not re.match(r'^[a-zA-Z0-9_\-]+\.xlsx$', filename):
                self.logger.error("Invalid filename detected")
                return
            self.workbook.save(filename)
            self.logger.info(f"Data saved to {filename}")
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {str(e)}")

    async def errback_close_page(self, failure):
        try:
            page = failure.request.meta.get("playwright_page")
            if page:
                await page.close()
            self.logger.error(f"Request failed: {failure.request.url} - {failure.value}")
        except Exception as e:
            self.logger.error(f"Error in errback: {str(e)}")