import scrapy
from scrapy_playwright.page import PageMethod
import random
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re
import json
import asyncio
from playwright.async_api import async_playwright
from fake_useragent import UserAgent


class AdvancedAmazonSpider(scrapy.Spider):
    name = "advanced_amazon_spider"
    allowed_domains = ["amazon.com"]
    
    custom_settings = {
        "PLAYWRIGHT_BROWSER_TYPE": "chromium",
        "PLAYWRIGHT_DEFAULT_NAVIGATION_TIMEOUT": 60000,
        "PLAYWRIGHT_LAUNCH_OPTIONS": {
            "headless": True,
            "timeout": 30000,
            "args": [
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--disable-web-security",
                "--disable-features=VizDisplayCompositor",
                "--disable-setuid-sandbox",
                "--disable-accelerated-2d-canvas",
                "--no-first-run",
                "--no-zygote",
                "--disable-gpu",
                "--disable-blink-features=AutomationControlled",
                "--disable-extensions",
                "--disable-extensions-except=",
                "--disable-plugins-discovery",
                "--enable-automation",
                "--start-maximized",
                "--disable-background-tasks",
                "--disable-backgrounding-occluded-windows",
                "--disable-renderer-backgrounding",
                "--no-default-browser-check",
                "--disable-default-apps",
                "--disable-ipc-flooding-protection",
                "--disable-background-networking"
            ]
        },
        "CONCURRENT_REQUESTS": 1,  # 降低并发数以避免被检测
        "DOWNLOAD_DELAY": 10,  # 增加下载延迟
        "RANDOMIZE_DOWNLOAD_DELAY": 0.5  # 随机化下载延迟
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Amazon商品数据"
        self.worksheet.append([
            "商品ID", "商品名称", "价格", "销量", "评分", "评论数", 
            "品牌", "类别", "上架时间", "采集时间", "URL", "配送信息", "是否Prime"
        ])
        
        # 配置参数
        self.min_wait_time = 3
        self.max_wait_time = 8
        self.scroll_wait_time = 3000
        self.retry_times = 3
        
        # 初始化User-Agent生成器
        try:
            self.ua = UserAgent()
        except:
            # 如果fake-useragent不可用，使用预定义的列表
            self.user_agents = [
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0',
                'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15'
            ]
        
        # 代理池配置
        self.proxies = [
            # 配置真实代理服务器，示例如下（请替换为实际可用的代理）
            # "http://username:password@proxy_host:proxy_port",
        ]
        
        # 请求计数器（用于控制请求频率）
        self.request_count = 0
        self.max_requests_per_minute = 10

    def get_random_user_agent(self):
        """获取随机User-Agent"""
        try:
            return self.ua.random
        except:
            return random.choice(self.user_agents)

    def get_random_proxy(self):
        """获取随机代理"""
        if self.proxies:
            return random.choice(self.proxies)
        return None

    def start_requests(self):
        # 从命令行参数获取搜索关键词
        search_terms = getattr(self, 'search', 'laptop')
        pages = getattr(self, 'pages', 3)  # 默认爬取3页
        
        # 构建搜索URL
        base_urls = []
        for page in range(1, pages + 1):
            url = f"https://www.amazon.com/s?k={search_terms}&page={page}"
            base_urls.append(url)
        
        for url in base_urls:
            yield self.create_request(url)
    
    async def start(self):
        # 从命令行参数获取搜索关键词
        search_terms = getattr(self, 'search', 'laptop')
        pages = getattr(self, 'pages', 3)  # 默认爬取3页
        
        # 构建搜索URL
        base_urls = []
        for page in range(1, pages + 1):
            url = f"https://www.amazon.com/s?k={search_terms}&page={page}"
            base_urls.append(url)
        
        for url in base_urls:
            yield self.create_request(url)

    def create_request(self, url):
        """创建带反爬策略的请求"""
        self.request_count += 1
        
        # 请求元数据
        meta = {
            "playwright": True,
            "playwright_page_coroutines": [
                PageMethod("wait_for_load_state", "networkidle"),
                PageMethod("wait_for_selector", "[data-component-type='s-search-result']"),
                PageMethod("evaluate", "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"),
                PageMethod("evaluate", "navigator.chrome = { runtime: {} }"),
                PageMethod("evaluate", "Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] })"),
                PageMethod("evaluate", "Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] })"),
            ],
            "playwright_context_kwargs": {
                "user_agent": self.get_random_user_agent(),
                "java_script_enabled": True,
                "viewport": {
                    "width": random.randint(1280, 1920),
                    "height": random.randint(720, 1080)
                }
            }
        }
        
        # 如果有代理，添加代理配置
        proxy = self.get_random_proxy()
        if proxy:
            meta["playwright_context_kwargs"]["proxy"] = {
                "server": proxy
            }
        
        return scrapy.Request(
            url,
            meta=meta,
            callback=self.parse,
            errback=self.errback_close_page,
            headers={
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none",
                "Cache-Control": "max-age=0",
                "DNT": "1",
                "Referer": "https://www.google.com/",
            },
            dont_filter=False
        )

    async def parse(self, response):
        self.logger.info(f"Parsing page: {response.url}")

        try:
            # 模拟随机停留时间，模拟人类行为
            await asyncio.sleep(random.uniform(self.min_wait_time, self.max_wait_time))

            # 模拟页面滚动行为
            if response.meta.get("playwright_page"):
                page = response.meta["playwright_page"]
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight/3);")
                await asyncio.sleep(random.uniform(1, 3))
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight*2/3);")
                await asyncio.sleep(random.uniform(1, 3))
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                await asyncio.sleep(random.uniform(2, 4))

            # 获取所有商品项
            products = response.css("[data-component-type='s-search-result']")
            
            self.logger.info(f"Found {len(products)} products on page {response.url}")
            
            for product in products:
                try:
                    # 提取商品信息
                    product_id = product.css("::attr(data-asin)").get()
                    if not product_id:
                        continue
                        
                    # 跳过广告产品
                    if product.css(".s-sponsored-faceout-badge-wrapper").get():
                        continue
                        
                    # 商品名称
                    name = product.css("h2 a span::text").get()
                    if name:
                        name = name.strip()
                    
                    # 价格信息
                    price_whole = product.css(".a-price-whole::text").get()
                    price_fraction = product.css(".a-price-fraction::text").get()
                    if price_whole:
                        price = f"${price_whole}{price_fraction or ''}"
                    else:
                        # 尝试其他价格选择器
                        price = product.css(".a-offscreen::text").get()
                    
                    # 评分
                    rating = product.css("span[aria-label*='out of']::attr(aria-label)").get()
                    if not rating:
                        rating_match = product.css("i span::text").re(r'[\d.]+')
                        rating = rating_match[0] if rating_match else "N/A"
                    
                    # 评论数
                    review_count_text = product.css("span[aria-label*='ratings']::text").get()
                    if not review_count_text:
                        review_count_match = product.css("a[role='link'] span::text").re(r'\d+,\d+|\d+')
                        review_count_text = review_count_match[0] if review_count_match else "0"
                    
                    # 构建商品详情页URL
                    product_url = product.css("h2 a::attr(href)").get()
                    if product_url:
                        product_url = response.urljoin(product_url)
                    else:
                        product_url = "N/A"
                    
                    # 数据验证和清理
                    if product_id and name:
                        # 清理数据
                        cleaned_name = name.strip() if name else ""
                        cleaned_price = price.strip() if price else "N/A"
                        cleaned_rating = rating.strip() if rating else "N/A"
                        cleaned_review_count = ''.join(filter(str.isdigit, review_count_text)) if review_count_text else "0"

                        # 添加到工作表
                        row_data = [
                            product_id,
                            cleaned_name,
                            cleaned_price,
                            "N/A",  # 销量暂时设为N/A
                            cleaned_rating,
                            cleaned_review_count,
                            "N/A",  # 品牌需要从详情页获取
                            "N/A",  # 分类需要从面包屑获取
                            "N/A",  # 上架时间需要从详情页获取
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            product_url,
                            "N/A",  # 配送信息
                            "N/A"   # 是否Prime
                        ]
                        
                        self.worksheet.append(row_data)
                        
                        self.logger.info(f"Added product: {cleaned_name[:50]}... (ID: {product_id})")
                        
                        # 请求商品详情页以获取更多信息
                        yield self.create_product_detail_request(product_url, product_id)
                            
                except Exception as e:
                    self.logger.error(f"Error parsing product on {response.url}: {str(e)}")
                    continue

            # 处理分页 - 尝试获取下一页链接
            next_page = response.css('a[aria-label="Next page"]::attr(href)').get()
            if not next_page:
                # 尝试其他下一页选择器
                next_page = response.css('li.a-last a::attr(href)').get()
            
            if next_page:
                next_page_url = response.urljoin(next_page)
                yield self.create_request(next_page_url)
                
        except Exception as e:
            self.logger.error(f"Error parsing page {response.url}: {str(e)}")

    def create_product_detail_request(self, url, product_id):
        """创建商品详情页请求"""
        self.request_count += 1
        
        meta = {
            "playwright": True,
            "playwright_page_coroutines": [
                PageMethod("wait_for_load_state", "networkidle"),
                PageMethod("wait_for_selector", "#productTitle"),
                PageMethod("evaluate", "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"),
                PageMethod("evaluate", "navigator.chrome = { runtime: {} }"),
                PageMethod("evaluate", "Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] })"),
                PageMethod("evaluate", "Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] })"),
            ],
            "playwright_context_kwargs": {
                "user_agent": self.get_random_user_agent(),
                "java_script_enabled": True,
                "viewport": {
                    "width": random.randint(1280, 1920),
                    "height": random.randint(720, 1080)
                }
            },
            "product_id": product_id
        }
        
        # 如果有代理，添加代理配置
        proxy = self.get_random_proxy()
        if proxy:
            meta["playwright_context_kwargs"]["proxy"] = {
                "server": proxy
            }
        
        return scrapy.Request(
            url=url,
            meta=meta,
            callback=self.parse_product_detail,
            errback=self.errback_close_page,
            headers={
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
                "Referer": "https://www.amazon.com/",
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none",
                "Cache-Control": "max-age=0",
                "DNT": "1"
            },
            dont_filter=True  # 不过滤重复URL，因为这是详情页
        )

    async def parse_product_detail(self, response):
        """解析商品详情页"""
        product_id = response.meta.get("product_id")
        self.logger.info(f"Parsing product detail: {response.url}")
        
        try:
            # 模拟人类行为 - 随机等待和页面交互
            await asyncio.sleep(random.uniform(2, 5))
            
            # 模拟页面滚动
            if response.meta.get("playwright_page"):
                page = response.meta["playwright_page"]
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight/4);")
                await asyncio.sleep(random.uniform(1, 2))
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight/2);")
                await asyncio.sleep(random.uniform(1, 2))
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                await asyncio.sleep(random.uniform(2, 3))
            
            # 查找对应的商品行并更新详细信息
            for row in self.worksheet.iter_rows(min_row=2, max_row=self.worksheet.max_row):
                if row[0].value == product_id:  # 商品ID匹配
                    # 提取品牌
                    brand = response.css("#bylineInfo::text").get()
                    if not brand:
                        brand = response.css("#brand ::text").get()
                    if not brand:
                        # 使用更兼容的选择器，避免使用contains函数
                        brand_elements = response.css("#bylineInfo-container a, #bylineInfo a")
                        if brand_elements:
                            brand = brand_elements[0].css("::text").get()
                        else:
                            brand = "N/A"
                    
                    # 提取分类
                    category = "N/A"
                    breadcrumb_items = response.css("#wayfinding-breadcrumbs_feature_div ul li span a::text").getall()
                    if breadcrumb_items:
                        category = " > ".join([item.strip() for item in breadcrumb_items if item.strip()])
                    
                    # 提取上架时间
                    date_info = response.css("#productDetails_detailBullets_sections1 td:contains('Date First Available') + td::text").get()
                    if not date_info:
                        # 使用更兼容的选择器，避免使用contains函数
                        date_labels = response.css("div.content ul li span::text").getall()
                        for i, label in enumerate(date_labels):
                            if label and 'Date First Available' in label:
                                date_info = response.css(f"div.content ul li span::text")[i+1].get()
                                break
                    
                    if not date_info:
                        # 尝试其他选择器
                        date_info = response.css("#detailBullets_feature_div ul li span[dir='auto']::text").re(r'Date First Available.*?(\w+ \d+, \d{4}|\w+ \d{4})')
                        if date_info:
                            date_info = date_info[0]
                    
                    # 提取配送信息
                    shipping_info = response.css("#deliveryMessageMirId::text").get()
                    if not shipping_info:
                        shipping_info = response.css("#exportsBuyBox_feature_div span::text").get()
                    if not shipping_info:
                        shipping_info = "N/A"
                    
                    # 检查是否Prime商品
                    is_prime = "Yes" if response.css(".a-icon-prime").get() else "No"
                    
                    # 提取销量排名（如果可用）
                    bestseller_badge = response.css("span.badge-text::text").get()
                    
                    # 更新行数据
                    row[6].value = brand.strip() if brand else "N/A"
                    row[7].value = category
                    row[8].value = date_info.strip() if date_info else "N/A"
                    row[11].value = shipping_info.strip() if shipping_info else "N/A"
                    row[12].value = is_prime
                    
                    self.logger.info(f"Updated product details for ID: {product_id}")
                    break

            # 滚动页面以加载更多评论（如果需要）
            if response.meta.get("playwright_page"):
                await response.meta["playwright_page"].evaluate("window.scrollTo(0, document.body.scrollHeight);")
                await asyncio.sleep(2)
            
        except Exception as e:
            self.logger.error(f"Error parsing product detail {response.url}: {str(e)}")

    def close(self, reason):
        try:
            # 保存Excel文件，防止路径遍历
            current_time = datetime.now().strftime("%Y%m%d%H%M%S")
            filename = f"advanced_amazon_data_{current_time}.xlsx"
            # 验证文件名安全
            if not re.match(r'^[a-zA-Z0-9_\-]+\.xlsx$', filename):
                self.logger.error("Invalid filename detected")
                return
            self.workbook.save(filename)
            self.logger.info(f"Data saved to {filename}")
            
            # 统计信息
            total_products = self.worksheet.max_row - 1
            self.logger.info(f"Total products collected: {total_products}")
            
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {str(e)}")

    async def errback_close_page(self, failure):
        try:
            page = failure.request.meta.get("playwright_page")
            if page:
                await page.close()
            self.logger.error(f"Request failed: {failure.request.url} - {failure.value}")
        except Exception as e:
            self.logger.error(f"Error in errback: {str(e)}")